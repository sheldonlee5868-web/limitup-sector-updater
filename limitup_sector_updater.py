#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
按 28 个行业板块自动统计每日涨停数，并将结果写入 Excel。

主要功能
--------
1. 读取最新的 TDX_Industry_merged.xlsx，生成“股票代码 -> 行业板块”映射。
2. 从东方财富涨停股池中抓取指定交易日的涨停股票，按 28 个板块统计涨停数。
3. 自动剔除 ST / *ST / SST 股票。
4. 自动提取二板、三板及以上的连板股名称，并写入对应板块单元格。
5. 统计上证指数(000001) + 深证成指(399001)的总成交额，以及相对前一交易日的变化。
6. 支持：
   - 增量更新（推荐，日常使用）
   - 指定日期区间重算（保留接口；过久远日期可能因公开数据源限制而失败）
   - 仅刷新 Dashboard 图表页
7. 输出 Excel Dashboard（折线图 + 最近 60 个交易日热力表）。

依赖安装
--------
pip install akshare pandas openpyxl matplotlib

示例
----
1) 日常增量更新（推荐）
python limitup_sector_updater.py ^
  --industry TDX_Industry_merged.xlsx ^
  --stats "2026-01 涨停板块统计（涨停数大于3标红加粗）.xlsx" ^
  --output "涨停板块统计_自动更新.xlsx"

2) 指定区间重算（仅保留接口；如果数据源已不支持旧日期，会自动报错提示）
python limitup_sector_updater.py ^
  --mode rebuild ^
  --start-date 2026-02-01 ^
  --end-date 2026-02-28 ^
  --industry TDX_Industry_merged.xlsx ^
  --stats "2026-01 涨停板块统计（涨停数大于3标红加粗）.xlsx" ^
  --output "涨停板块统计_重算.xlsx"

3) 仅刷新 Dashboard 图表
python limitup_sector_updater.py --mode dashboard --stats "涨停板块统计_自动更新.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import math
import os
import random
import re
import sys
import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import akshare as ak
except ImportError:  # pragma: no cover
    ak = None


DATE_ROW = 1
SUMMARY_ROW = 2
SECTOR_START_ROW = 3
FIRST_DATA_COL = 2
DEFAULT_RECENT_LOOKBACK_DAYS = 30
SH_INDEX = "000001"  # 上证指数
SZ_INDEX = "399001"  # 深证成指
PROXY_ENV_KEYS = [
    "HTTP_PROXY",
    "HTTPS_PROXY",
    "ALL_PROXY",
    "http_proxy",
    "https_proxy",
    "all_proxy",
]


@dataclass
class DailyResult:
    trade_date: str
    total_limit_up: int
    turnover_total: float  # 元
    turnover_delta: Optional[float]  # 元
    market_tag: str
    sector_counts: Dict[str, int]
    sector_lb_names: Dict[str, List[str]]


@dataclass
class RunSummary:
    mode: str
    start_date: str
    end_date: str
    trade_dates: List[str]
    success_dates: List[str]
    failed_dates: List[str]
    output_path: Path
    log_path: Optional[Path]


def setup_logging(verbose: bool = False, log_file: Optional[Path] = None) -> None:
    handlers = [logging.StreamHandler(sys.stdout)]
    if log_file:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))

    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="[%(asctime)s] %(levelname)s - %(message)s",
        datefmt="%H:%M:%S",
        handlers=handlers,
        force=True,
    )

    if log_file:
        logging.info("日志文件：%s", log_file)


def ensure_akshare() -> None:
    if ak is None:
        raise SystemExit(
            "未检测到 akshare，请先安装：\n"
            "pip install akshare pandas openpyxl matplotlib"
        )


def normalize_code(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().replace("'", "")
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d{1,6})", text)
    if not match:
        return None
    return match.group(1).zfill(6)


def normalize_sector_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"\(\d+\)$", "", text)
    return text


def format_sector_label(sector: str, stock_count: int) -> str:
    return f"{sector}（{stock_count}）"


def is_st_stock(name: object) -> bool:
    if name is None:
        return False
    text = str(name).strip().upper().replace(" ", "")
    return text.startswith("ST") or text.startswith("*ST") or text.startswith("SST")


def parse_date_cell(value: object) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    dt = pd.to_datetime(text, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.strftime("%Y-%m-%d")


def parse_leading_int(value: object) -> Optional[int]:
    if value in (None, ""):
        return None
    match = re.match(r"\s*(\d+)", str(value))
    return int(match.group(1)) if match else None


def previous_calendar_day(ymd: str, days: int = 15) -> str:
    dt = datetime.strptime(ymd, "%Y-%m-%d") - timedelta(days=days)
    return dt.strftime("%Y-%m-%d")


def ymd_to_compact(ymd: str) -> str:
    return ymd.replace("-", "")


def compact_to_ymd(text: str) -> str:
    return datetime.strptime(text, "%Y%m%d").strftime("%Y-%m-%d")


def compute_retry_wait(sleep_seconds: float, attempt: int, max_wait: float = 12.0) -> float:
    base = max(sleep_seconds, 0.2) * (1.8 ** max(attempt - 1, 0))
    jitter = random.uniform(0, min(0.5, sleep_seconds if sleep_seconds > 0 else 0.5))
    return min(max_wait, base + jitter)


def retry_call(func, *args, retries: int = 3, sleep_seconds: float = 1.0, **kwargs):
    last_error = None
    retry_name = getattr(func, "__name__", str(func))
    for attempt in range(1, retries + 1):
        try:
            return func(*args, **kwargs)
        except Exception as exc:  # pragma: no cover
            last_error = exc
            if attempt < retries:
                wait = compute_retry_wait(sleep_seconds, attempt)
                logging.warning(
                    "调用 %s 失败，第 %s/%s 次：%s；%.1f 秒后重试",
                    retry_name,
                    attempt,
                    retries,
                    exc,
                    wait,
                )
                time.sleep(wait)
            else:
                logging.warning("调用 %s 失败，第 %s/%s 次：%s", retry_name, attempt, retries, exc)
    raise last_error


def is_proxy_error(exc: Exception) -> bool:
    text = f"{type(exc).__name__}: {exc}"
    return "ProxyError" in text or "Unable to connect to proxy" in text


def disable_env_proxies() -> List[str]:
    removed_keys: List[str] = []
    for key in PROXY_ENV_KEYS:
        if key in os.environ:
            os.environ.pop(key, None)
            removed_keys.append(key)
    return removed_keys


def configure_proxy_mode(proxy_mode: str) -> None:
    if proxy_mode == "direct":
        removed = disable_env_proxies()
        if removed:
            logging.info("已按直连模式禁用代理环境变量：%s", ", ".join(removed))
        else:
            logging.info("已按直连模式运行，当前未发现代理环境变量。")
    else:
        active = [key for key in PROXY_ENV_KEYS if os.environ.get(key)]
        if active:
            logging.info("检测到代理环境变量：%s", ", ".join(active))


def call_data_api(func, *args, proxy_mode: str = "auto", retries: int = 3, sleep_seconds: float = 1.0, **kwargs):
    try:
        return retry_call(func, *args, retries=retries, sleep_seconds=sleep_seconds, **kwargs)
    except Exception as exc:
        if proxy_mode == "auto" and is_proxy_error(exc):
            removed = disable_env_proxies()
            if removed:
                logging.warning("检测到代理异常，已自动切换为直连模式并重试。已禁用：%s", ", ".join(removed))
            else:
                logging.warning("检测到代理异常，已自动重试直连模式。")
            return retry_call(func, *args, retries=retries, sleep_seconds=sleep_seconds, **kwargs)
        raise


def load_industry_mapping(industry_path: Path) -> Tuple[List[str], Dict[str, List[str]], Dict[str, str]]:
    wb = openpyxl.load_workbook(industry_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    sector_order: List[str] = []
    sector_to_codes: Dict[str, List[str]] = {}
    code_to_sector: Dict[str, str] = {}

    for col in range(1, ws.max_column + 1, 2):
        sector = ws.cell(1, col).value
        if sector in (None, ""):
            continue
        sector_name = str(sector).strip()
        sector_order.append(sector_name)
        sector_to_codes[sector_name] = []
        for row in range(2, ws.max_row + 1):
            code = normalize_code(ws.cell(row, col).value)
            if not code:
                continue
            sector_to_codes[sector_name].append(code)
            code_to_sector[code] = sector_name

    if len(sector_order) != 28:
        logging.warning("检测到行业板块数量为 %s（预期通常为 28）。程序仍将按当前文件内容运行。", len(sector_order))

    logging.info("已读取行业板块 %s 个，股票 %s 只。", len(sector_order), len(code_to_sector))
    return sector_order, sector_to_codes, code_to_sector


def init_new_workbook(sector_order: List[str], sector_to_codes: Dict[str, List[str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 16

    blue_font = Font(color="FF0070C0")
    row_label_font = Font(color="FF0000FF")
    normal_font = Font(color="FF000000")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(1, SECTOR_START_ROW + len(sector_order)):
        ws.row_dimensions[row].height = 30

    ws.cell(DATE_ROW, 1).value = None
    ws.cell(SUMMARY_ROW, 1).value = None

    for row_index, sector in enumerate(sector_order, start=SECTOR_START_ROW):
        c = ws.cell(row_index, 1)
        c.value = format_sector_label(sector, len(sector_to_codes[sector]))
        c.font = copy(row_label_font)
        c.alignment = copy(center_wrap)
        c.number_format = "@"

    for row in (DATE_ROW, SUMMARY_ROW):
        c = ws.cell(row, FIRST_DATA_COL)
        c.font = copy(blue_font)
        c.alignment = copy(center_wrap)
        c.number_format = "yyyy-mm-dd" if row == DATE_ROW else "@"
    for row in range(SECTOR_START_ROW, SECTOR_START_ROW + len(sector_order)):
        c = ws.cell(row, FIRST_DATA_COL)
        c.font = copy(normal_font)
        c.alignment = copy(center_wrap)
        c.number_format = "@"

    return wb


def load_or_create_workbook(stats_path: Path, sector_order: List[str], sector_to_codes: Dict[str, List[str]]) -> Tuple[Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if stats_path.exists():
        wb = openpyxl.load_workbook(stats_path)
        ws = wb[wb.sheetnames[0]]
        logging.info("已载入统计工作簿：%s", stats_path)
    else:
        wb = init_new_workbook(sector_order, sector_to_codes)
        ws = wb[wb.sheetnames[0]]
        logging.info("统计工作簿不存在，已创建新工作簿模板。")
    return wb, ws


def build_sector_row_map(ws, sector_order: List[str], sector_to_codes: Dict[str, List[str]]) -> Dict[str, int]:
    existing: Dict[str, int] = {}
    for row in range(SECTOR_START_ROW, max(ws.max_row, SECTOR_START_ROW + len(sector_order) - 1) + 1):
        label = ws.cell(row, 1).value
        key = normalize_sector_name(label)
        if key:
            existing[key] = row

    row_map: Dict[str, int] = {}
    used_rows = set()
    next_row = SECTOR_START_ROW
    for sector in sector_order:
        key = normalize_sector_name(sector)
        if key in existing:
            row = existing[key]
        else:
            while next_row in used_rows or ws.cell(next_row, 1).value not in (None, ""):
                next_row += 1
            row = next_row
        used_rows.add(row)
        row_map[sector] = row
        label_cell = ws.cell(row, 1)
        label_cell.value = format_sector_label(sector, len(sector_to_codes[sector]))
        label_cell.number_format = "@"
        if not label_cell.alignment or label_cell.alignment == Alignment():
            label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if not label_cell.font or label_cell.font == Font():
            label_cell.font = Font(color="FF0000FF")
        ws.row_dimensions[row].height = ws.row_dimensions[row].height or 30
    return row_map


def scan_existing_dates(ws) -> Tuple[Dict[str, int], List[Tuple[int, str]]]:
    raw_items: List[Tuple[int, str]] = []
    for col in range(FIRST_DATA_COL, ws.max_column + 1):
        date_str = parse_date_cell(ws.cell(DATE_ROW, col).value)
        if date_str:
            raw_items.append((col, date_str))

    parsed_dates = [datetime.strptime(date_str, "%Y-%m-%d") for _, date_str in raw_items]
    outlier_idx = set()

    for idx in range(1, len(parsed_dates) - 1):
        prev_dt = parsed_dates[idx - 1]
        curr_dt = parsed_dates[idx]
        next_dt = parsed_dates[idx + 1]

        # 只剔除“夹在两个正常日期之间的明显录入异常值”，避免误删其左右两侧的正常日期。
        is_low_outlier = curr_dt < prev_dt and next_dt >= prev_dt
        is_high_outlier = curr_dt > next_dt and prev_dt <= next_dt
        if is_low_outlier or is_high_outlier:
            outlier_idx.add(idx)

    date_map: Dict[str, int] = {}
    outliers: List[Tuple[int, str]] = []
    for idx, (col, date_str) in enumerate(raw_items):
        if idx in outlier_idx:
            outliers.append((col, date_str))
            continue
        date_map[date_str] = col
    return date_map, outliers


def get_existing_date_map(ws, warn: bool = False) -> Dict[str, int]:
    date_map, outliers = scan_existing_dates(ws)
    if warn:
        for col, date_str in outliers:
            logging.warning("发现疑似异常日期列：第 %s 列 = %s，已忽略该列。", col, date_str)
    return date_map


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def ensure_date_column(ws, trade_date: str, sector_row_count: int, existing_date_map: Optional[Dict[str, int]] = None) -> int:
    existing = existing_date_map if existing_date_map is not None else get_existing_date_map(ws)
    if trade_date in existing:
        return existing[trade_date]

    new_col = max(ws.max_column + 1, FIRST_DATA_COL)
    ref_col = new_col - 1 if new_col > FIRST_DATA_COL else None

    if ref_col is not None and ref_col >= FIRST_DATA_COL:
        max_row = max(ws.max_row, SECTOR_START_ROW + sector_row_count - 1)
        for row in range(1, max_row + 1):
            copy_cell_style(ws.cell(row, ref_col), ws.cell(row, new_col))
        ref_letter = get_column_letter(ref_col)
        new_letter = get_column_letter(new_col)
        ws.column_dimensions[new_letter].width = ws.column_dimensions[ref_letter].width or 13
    else:
        ws.column_dimensions[get_column_letter(new_col)].width = 13

    if existing_date_map is not None:
        existing_date_map[trade_date] = new_col
    return new_col


def build_requests_session(proxy_mode: str = "auto") -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json,text/plain,*/*",
        "Referer": "https://quote.eastmoney.com/",
        "Origin": "https://quote.eastmoney.com",
        "Connection": "keep-alive",
    })
    if proxy_mode == "direct":
        session.trust_env = False
        session.proxies.update({"http": None, "https": None})
    return session


def fetch_index_history_by_eastmoney_direct(
    symbol: str,
    start_date: str,
    end_date: str,
    proxy_mode: str = "auto",
    retries: int = 3,
    sleep_seconds: float = 1.0,
) -> pd.DataFrame:
    secid_map = {
        SH_INDEX: "1.000001",  # 上证指数
        SZ_INDEX: "0.399001",  # 深证成指
    }
    secid = secid_map.get(symbol)
    if not secid:
        raise ValueError(f"暂不支持的指数代码：{symbol}")

    url = "https://push2his.eastmoney.com/api/qt/stock/kline/get"
    params = {
        "secid": secid,
        "fields1": "f1,f2,f3,f4,f5,f6",
        "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61",
        "klt": "101",
        "fqt": "0",
        "beg": ymd_to_compact(start_date),
        "end": ymd_to_compact(end_date),
        "ut": "fa5fd1943c7b386f172d6893dbfba10b",
    }

    last_error = None
    for attempt in range(1, retries + 1):
        try:
            session = build_requests_session(proxy_mode=proxy_mode)
            resp = session.get(url, params=params, timeout=20)
            resp.raise_for_status()
            data = resp.json()
            if not isinstance(data, dict) or not data.get("data"):
                raise RuntimeError(f"东财指数接口返回为空：{symbol}")
            klines = data["data"].get("klines") or []
            if not klines:
                raise RuntimeError(f"东财指数接口无K线数据：{symbol}")
            rows = []
            for item in klines:
                parts = str(item).split(",")
                if len(parts) < 10:
                    continue
                rows.append({
                    "日期": parts[0],
                    "开盘": pd.to_numeric(parts[1], errors="coerce"),
                    "收盘": pd.to_numeric(parts[2], errors="coerce"),
                    "最高": pd.to_numeric(parts[3], errors="coerce"),
                    "最低": pd.to_numeric(parts[4], errors="coerce"),
                    "成交量": pd.to_numeric(parts[5], errors="coerce"),
                    "成交额": pd.to_numeric(parts[6], errors="coerce"),
                    "振幅": pd.to_numeric(parts[7], errors="coerce"),
                    "涨跌幅": pd.to_numeric(parts[8], errors="coerce"),
                    "涨跌额": pd.to_numeric(parts[9], errors="coerce"),
                    "换手率": pd.to_numeric(parts[10], errors="coerce") if len(parts) > 10 else None,
                })
            df = pd.DataFrame(rows)
            if df.empty:
                raise RuntimeError(f"东财指数接口解析后为空：{symbol}")
            return df
        except Exception as exc:  # pragma: no cover
            last_error = exc
            logging.warning(
                "直连东财指数接口失败(%s)，第 %s/%s 次：%s",
                symbol, attempt, retries, exc
            )
            if attempt < retries:
                time.sleep(sleep_seconds)
    raise last_error


def fetch_index_history(start_date: str, end_date: str, retries: int, sleep_seconds: float, proxy_mode: str = "auto") -> pd.DataFrame:
    logging.info("抓取指数历史数据：%s ~ %s", start_date, end_date)

    df_sh = None
    df_sz = None

    try:
        df_sh = fetch_index_history_by_eastmoney_direct(
            SH_INDEX, start_date, end_date, proxy_mode=proxy_mode, retries=retries, sleep_seconds=sleep_seconds
        )
        df_sz = fetch_index_history_by_eastmoney_direct(
            SZ_INDEX, start_date, end_date, proxy_mode=proxy_mode, retries=retries, sleep_seconds=sleep_seconds
        )
        logging.info("指数历史数据已切换为东财直连模式。")
    except Exception as direct_exc:
        logging.warning("东财直连指数接口失败，回退到 akshare.index_zh_a_hist：%s", direct_exc)
        df_sh = call_data_api(
            ak.index_zh_a_hist,
            symbol=SH_INDEX,
            period="daily",
            start_date=ymd_to_compact(start_date),
            end_date=ymd_to_compact(end_date),
            proxy_mode=proxy_mode,
            retries=retries,
            sleep_seconds=sleep_seconds,
        )
        df_sz = call_data_api(
            ak.index_zh_a_hist,
            symbol=SZ_INDEX,
            period="daily",
            start_date=ymd_to_compact(start_date),
            end_date=ymd_to_compact(end_date),
            proxy_mode=proxy_mode,
            retries=retries,
            sleep_seconds=sleep_seconds,
        )

    for required in ("日期", "成交额", "涨跌幅"):
        if required not in df_sh.columns or required not in df_sz.columns:
            raise RuntimeError("指数接口返回字段异常，请检查 akshare 版本或数据源是否变更。")

    sh = df_sh[["日期", "成交额", "涨跌幅"]].copy()
    sh.columns = ["日期", "上证成交额", "上证涨跌幅"]
    sz = df_sz[["日期", "成交额", "涨跌幅"]].copy()
    sz.columns = ["日期", "深证成交额", "深证涨跌幅"]

    merged = pd.merge(sh, sz, on="日期", how="inner")
    merged["日期"] = pd.to_datetime(merged["日期"])
    merged = merged.sort_values("日期").reset_index(drop=True)
    merged["总成交额"] = pd.to_numeric(merged["上证成交额"], errors="coerce").astype(float) + pd.to_numeric(merged["深证成交额"], errors="coerce").astype(float)
    merged["总成交额变化"] = merged["总成交额"].diff()
    merged["日期字符串"] = merged["日期"].dt.strftime("%Y-%m-%d")
    merged["市场标签"] = merged.apply(classify_market_tag, axis=1)
    return merged


def classify_market_tag(row: pd.Series) -> str:
    try:
        sh_pct = float(row["上证涨跌幅"])
        sz_pct = float(row["深证涨跌幅"])
    except Exception:
        return ""

    if sh_pct > 0 and sz_pct > 0:
        return "上涨"
    if sh_pct < 0 and sz_pct < 0:
        return "下跌"
    return "震荡"


def fetch_limit_up_pool(trade_date: str, retries: int, sleep_seconds: float, proxy_mode: str = "auto") -> pd.DataFrame:
    compact = ymd_to_compact(trade_date)
    logging.info("抓取涨停股池：%s", trade_date)
    df = call_data_api(
        ak.stock_zt_pool_em,
        date=compact,
        proxy_mode=proxy_mode,
        retries=retries,
        sleep_seconds=sleep_seconds,
    )
    if df is None or df.empty:
        raise RuntimeError(f"涨停股池返回空数据：{trade_date}")

    if "代码" not in df.columns or "名称" not in df.columns:
        raise RuntimeError(f"涨停股池返回字段异常：{trade_date}")

    df = df.copy()
    df["代码"] = df["代码"].map(normalize_code)
    df["名称"] = df["名称"].astype(str).str.strip()
    if "连板数" not in df.columns:
        df["连板数"] = 1
    df["连板数"] = pd.to_numeric(df["连板数"], errors="coerce").fillna(1).astype(int)
    df = df.dropna(subset=["代码"]).drop_duplicates(subset=["代码"], keep="first")
    return df


def build_daily_result(
    trade_date: str,
    pool_df: pd.DataFrame,
    index_row: pd.Series,
    sector_order: List[str],
    code_to_sector: Dict[str, str],
) -> DailyResult:
    sector_counts = {sector: 0 for sector in sector_order}
    sector_lb_names: Dict[str, List[str]] = {sector: [] for sector in sector_order}

    if not pool_df.empty:
        filtered = pool_df.copy()
        filtered = filtered[filtered["代码"].isin(code_to_sector)]
        filtered = filtered[~filtered["名称"].map(is_st_stock)]
        filtered["板块"] = filtered["代码"].map(code_to_sector)

        for _, row in filtered.iterrows():
            sector = row["板块"]
            sector_counts[sector] += 1
            if int(row.get("连板数", 1)) >= 2:
                sector_lb_names[sector].append(str(row["名称"]))

        for sector, names in sector_lb_names.items():
            if names:
                sector_lb_names[sector] = sorted(set(names))
    else:
        filtered = pool_df

    turnover_total = float(index_row["总成交额"])
    turnover_delta = index_row["总成交额变化"]
    turnover_delta = None if pd.isna(turnover_delta) else float(turnover_delta)

    return DailyResult(
        trade_date=trade_date,
        total_limit_up=int(len(filtered)),
        turnover_total=turnover_total,
        turnover_delta=turnover_delta,
        market_tag=str(index_row.get("市场标签", "") or ""),
        sector_counts=sector_counts,
        sector_lb_names=sector_lb_names,
    )


def format_turnover_line(delta: Optional[float], market_tag: str) -> str:
    if delta is None or abs(delta) < 1:
        base = "平量"
    elif delta > 0:
        base = f"放量{abs(delta) / 1e8:.0f}亿"
    else:
        base = f"缩量{abs(delta) / 1e8:.0f}亿"
    return f"{base}{market_tag}" if market_tag else base


def build_summary_cell_text(result: DailyResult) -> str:
    lines = [str(result.total_limit_up)]
    lines.append(format_turnover_line(result.turnover_delta, result.market_tag))
    lines.append(f"成交{result.turnover_total / 1e8:.0f}亿")
    return "\n".join(lines)


def build_sector_cell_text(count: int, lb_names: List[str]) -> Optional[str]:
    if count <= 0:
        return None
    if lb_names:
        return f"{count}\n{'/'.join(lb_names)}"
    return str(count)


def apply_value_style(cell, is_alert: bool) -> None:
    cell.number_format = "@"
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if is_alert:
        cell.font = Font(color="FFFF0000", bold=True)
    else:
        cell.font = Font(color="FF000000", bold=False)


def write_daily_result(ws, col: int, result: DailyResult, sector_order: List[str], row_map: Dict[str, int]) -> None:
    date_cell = ws.cell(DATE_ROW, col)
    summary_cell = ws.cell(SUMMARY_ROW, col)

    date_cell.value = datetime.strptime(result.trade_date, "%Y-%m-%d")
    date_cell.number_format = "yyyy-mm-dd"
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.font = Font(color="FF0070C0")

    summary_cell.value = build_summary_cell_text(result)
    summary_cell.number_format = "@"
    summary_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    summary_cell.font = Font(color="FF0070C0")

    for sector in sector_order:
        row = row_map[sector]
        count = result.sector_counts.get(sector, 0)
        lb_names = result.sector_lb_names.get(sector, [])
        cell = ws.cell(row, col)
        cell.value = build_sector_cell_text(count, lb_names)
        apply_value_style(cell, is_alert=(count > 3))

    # 统一行高
    max_row = max(row_map.values())
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = ws.row_dimensions[row].height or 30


def collect_trade_dates(index_df: pd.DataFrame, start_date: str, end_date: str) -> List[str]:
    mask = (index_df["日期字符串"] >= start_date) & (index_df["日期字符串"] <= end_date)
    return index_df.loc[mask, "日期字符串"].tolist()


def parse_summary_metrics(value: object) -> Tuple[Optional[int], Optional[float], Optional[float]]:
    if value in (None, ""):
        return None, None, None
    text = str(value)
    total_count = parse_leading_int(text)

    turnover_total = None
    turnover_delta = None

    m_total = re.search(r"成交\s*([0-9.]+)亿", text)
    if m_total:
        turnover_total = float(m_total.group(1))

    m_delta = re.search(r"(放量|缩量)\s*([0-9.]+)亿", text)
    if m_delta:
        delta = float(m_delta.group(2))
        turnover_delta = delta if m_delta.group(1) == "放量" else -delta
    elif "平量" in text:
        turnover_delta = 0.0

    return total_count, turnover_total, turnover_delta


def parse_market_tag(value: object) -> str:
    if value in (None, ""):
        return ""
    lines = str(value).splitlines()
    if len(lines) < 2:
        return ""
    line2 = lines[1]
    for tag in ("上涨", "下跌", "震荡"):
        if line2.endswith(tag):
            return tag
    return ""


def apply_cell_border(cell, color: str = "FFD9E2F3") -> None:
    side = Side(style="thin", color=color)
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def style_card_block(ws, start_col: int, end_col: int, title_row: int, value_row: int, title: str, value: object, fill_color: str) -> None:
    ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
    ws.merge_cells(start_row=value_row, start_column=start_col, end_row=value_row, end_column=end_col)
    title_cell = ws.cell(title_row, start_col)
    value_cell = ws.cell(value_row, start_col)

    title_cell.value = title
    value_cell.value = value

    title_fill = PatternFill("solid", fgColor=fill_color)
    value_fill = PatternFill("solid", fgColor="FFF8FBFF")

    title_cell.fill = title_fill
    value_cell.fill = value_fill
    title_cell.font = Font(color="FFFFFFFF", bold=True, size=11)
    value_cell.font = Font(color="FF1F1F1F", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in (title_row, value_row):
        for col in range(start_col, end_col + 1):
            apply_cell_border(ws.cell(row, col), color="FFB7C9E2")


def build_heat_fill(value: int) -> Tuple[str, str]:
    if value <= 0:
        return "FFF3F6FB", "FF7A869A"
    if value == 1:
        return "FFD9F0FF", "FF0F3B5F"
    if value == 2:
        return "FFADE8F4", "FF083D4F"
    if value == 3:
        return "FF95D5B2", "FF10451D"
    if value == 4:
        return "FFFFE08A", "FF6B4F00"
    if value == 5:
        return "FFFFC971", "FF6A2E00"
    if value == 6:
        return "FFFF9B71", "FF6A1B00"
    return "FFF94144", "FFFFFFFF"


def beautify_chart(chart, style_id: int, y_title: str, x_title: str) -> None:
    chart.style = style_id
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.legend = None
    chart.height = 8
    chart.width = 16
    if getattr(chart, "x_axis", None) is not None:
        chart.x_axis.tickLblPos = "low"
    if getattr(chart, "y_axis", None) is not None:
        chart.y_axis.majorGridlines = None


def write_heat_legend(ws, row: int, start_col: int) -> None:
    items = [
        ("0", "FFF3F6FB", "FF7A869A"),
        ("1", "FFD9F0FF", "FF0F3B5F"),
        ("2", "FFADE8F4", "FF083D4F"),
        ("3", "FF95D5B2", "FF10451D"),
        ("4", "FFFFE08A", "FF6B4F00"),
        ("5", "FFFFC971", "FF6A2E00"),
        ("6", "FFFF9B71", "FF6A1B00"),
        ("7+", "FFF94144", "FFFFFFFF"),
    ]
    ws.cell(row, start_col).value = "热度图例"
    ws.cell(row, start_col).font = Font(bold=True, color="FF1F1F1F")
    for idx, (label, fill_color, font_color) in enumerate(items, start=1):
        cell = ws.cell(row, start_col + idx)
        cell.value = label
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(color=font_color, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_cell_border(cell, color="FFE2E8F0")


def rebuild_dashboard(wb: Workbook, ws_main, sector_order: List[str], row_map: Dict[str, int]) -> None:
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B8"

    title_fill = PatternFill("solid", fgColor="FF173F5F")
    section_fill = PatternFill("solid", fgColor="FF1D4E89")
    header_fill = PatternFill("solid", fgColor="FFEAF2FF")
    stripe_fill = PatternFill("solid", fgColor="FFF8FBFF")
    strong_font = Font(bold=True, color="FF1F1F1F")
    white_font = Font(bold=True, color="FFFFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    date_map = get_existing_date_map(ws_main)
    sorted_items = sorted(date_map.items(), key=lambda x: x[0])

    ws.merge_cells("A1:F2")
    ws["A1"] = "涨停板块热力总览"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    subtitle = "暂无可展示数据"
    if sorted_items:
        subtitle = f"统计区间：{sorted_items[0][0]} ~ {sorted_items[-1][0]}｜最近 60 个交易日热力图"
    ws.merge_cells("A3:F3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(color="FF4A5568", italic=True)

    # 汇总表起始位置
    summary_header_row = 6
    summary_data_row = summary_header_row + 1

    latest_date = "-"
    latest_total = "-"
    latest_turnover = "-"
    hottest_sector = "-"
    hottest_count = "-"

    if sorted_items:
        latest_date, latest_col = sorted_items[-1]
        latest_summary = ws_main.cell(SUMMARY_ROW, latest_col).value
        latest_total_count, latest_turnover_total, _ = parse_summary_metrics(latest_summary)
        latest_total = latest_total_count if latest_total_count is not None else "-"
        latest_turnover = f"{latest_turnover_total:.0f} 亿" if latest_turnover_total is not None else "-"

        ranking_pairs = []
        for sector in sector_order:
            main_row = row_map[sector]
            count = parse_leading_int(ws_main.cell(main_row, latest_col).value) or 0
            ranking_pairs.append((sector, count))
        ranking_pairs.sort(key=lambda x: (-x[1], x[0]))
        if ranking_pairs:
            hottest_sector = ranking_pairs[0][0]
            hottest_count = ranking_pairs[0][1]

    style_card_block(ws, 7, 8, 1, 2, "最新交易日", latest_date, "FF1F78B4")
    style_card_block(ws, 9, 10, 1, 2, "总涨停数", latest_total, "FF33A02C")
    style_card_block(ws, 11, 12, 1, 2, "总成交额", latest_turnover, "FFFFA000")
    style_card_block(ws, 13, 14, 1, 2, "最热板块", hottest_sector, "FFE63946")
    style_card_block(ws, 15, 16, 1, 2, "板块峰值", hottest_count, "FF6A4C93")

    # 汇总表
    headers = ["日期", "总涨停数", "总成交额(亿)", "较前一日变化(亿)", "市场标签"]
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(summary_header_row, idx)
        cell.value = title
        cell.fill = section_fill
        cell.font = white_font
        cell.alignment = center
        apply_cell_border(cell, color="FFB7C9E2")

    for idx, (trade_date, col) in enumerate(sorted_items, start=summary_data_row):
        summary_value = ws_main.cell(SUMMARY_ROW, col).value
        total_count, turnover_total, turnover_delta = parse_summary_metrics(summary_value)
        market_tag = parse_market_tag(summary_value)

        ws.cell(idx, 1).value = datetime.strptime(trade_date, "%Y-%m-%d")
        ws.cell(idx, 1).number_format = "yyyy-mm-dd"
        ws.cell(idx, 2).value = total_count
        ws.cell(idx, 3).value = turnover_total
        ws.cell(idx, 4).value = turnover_delta
        ws.cell(idx, 5).value = market_tag

        row_fill = stripe_fill if (idx - summary_data_row) % 2 == 0 else None
        for col_idx in range(1, 6):
            cell = ws.cell(idx, col_idx)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center
            apply_cell_border(cell, color="FFE2E8F0")
            if col_idx == 5:
                cell.font = Font(
                    color=("FF0F9D58" if market_tag == "上涨" else "FFD93025" if market_tag == "下跌" else "FF5F6368"),
                    bold=True,
                )

    last_row = max(ws.max_row, summary_data_row)

    # 图表区
    if len(sorted_items) >= 1:
        cats = Reference(ws, min_col=1, min_row=summary_data_row, max_row=last_row)

        chart1 = LineChart()
        chart1.title = "每日总涨停数"
        chart1.add_data(Reference(ws, min_col=2, min_row=summary_header_row, max_row=last_row), titles_from_data=True)
        chart1.set_categories(cats)
        beautify_chart(chart1, style_id=13, y_title="家数", x_title="日期")
        ws.add_chart(chart1, "G6")

        chart2 = LineChart()
        chart2.title = "上证指数 + 深证成指总成交额"
        chart2.add_data(Reference(ws, min_col=3, min_row=summary_header_row, max_row=last_row), titles_from_data=True)
        chart2.set_categories(cats)
        beautify_chart(chart2, style_id=12, y_title="亿", x_title="日期")
        ws.add_chart(chart2, "G22")

        chart3 = BarChart()
        chart3.title = "较前一日成交额变化"
        chart3.add_data(Reference(ws, min_col=4, min_row=summary_header_row, max_row=last_row), titles_from_data=True)
        chart3.set_categories(cats)
        beautify_chart(chart3, style_id=11, y_title="亿", x_title="日期")
        chart3.dLbls = DataLabelList()
        chart3.dLbls.showVal = True
        ws.add_chart(chart3, "G38")

    # 最新交易日板块 TOP10
    rank_header_row = 6
    rank_start_row = 7
    rank_title_cell = ws.cell(rank_header_row, 18)
    rank_title_cell.value = "最新交易日板块热度 TOP10"
    rank_title_cell.fill = section_fill
    rank_title_cell.font = white_font
    rank_title_cell.alignment = center
    apply_cell_border(rank_title_cell, color="FFB7C9E2")
    ws.cell(rank_header_row, 19).fill = section_fill
    apply_cell_border(ws.cell(rank_header_row, 19), color="FFB7C9E2")
    ws.cell(rank_header_row, 19).font = white_font

    if sorted_items:
        latest_col = sorted_items[-1][1]
        ranked = []
        for sector in sector_order:
            main_row = row_map[sector]
            count = parse_leading_int(ws_main.cell(main_row, latest_col).value) or 0
            ranked.append((sector, count))
        ranked = sorted(ranked, key=lambda x: (-x[1], x[0]))[:10]

        ws.cell(rank_header_row, 18).value = "板块"
        ws.cell(rank_header_row, 19).value = "涨停数"
        for cell in (ws.cell(rank_header_row, 18), ws.cell(rank_header_row, 19)):
            cell.fill = section_fill
            cell.font = white_font
            cell.alignment = center
            apply_cell_border(cell, color="FFB7C9E2")

        for offset, (sector, count) in enumerate(ranked, start=rank_start_row):
            ws.cell(offset, 18).value = sector
            ws.cell(offset, 19).value = count
            for col_idx in (18, 19):
                cell = ws.cell(offset, col_idx)
                if (offset - rank_start_row) % 2 == 0:
                    cell.fill = stripe_fill
                cell.alignment = center
                apply_cell_border(cell, color="FFE2E8F0")

        if ranked:
            chart4 = BarChart()
            chart4.type = "bar"
            chart4.title = "最新交易日强势板块"
            chart4.add_data(Reference(ws, min_col=19, min_row=rank_header_row, max_row=rank_start_row + len(ranked) - 1), titles_from_data=True)
            chart4.set_categories(Reference(ws, min_col=18, min_row=rank_start_row, max_row=rank_start_row + len(ranked) - 1))
            beautify_chart(chart4, style_id=10, y_title="板块", x_title="涨停数")
            chart4.dLbls = DataLabelList()
            chart4.dLbls.showVal = True
            ws.add_chart(chart4, "Q22")

    # 最近 60 交易日板块热力图
    recent_items = sorted_items[-60:] if len(sorted_items) > 60 else sorted_items
    start_row = max(last_row + 5, 56)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(10, len(recent_items) + 3))
    title_cell = ws.cell(start_row, 1)
    title_cell.value = "最近 60 个交易日板块热力图"
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFFFF", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    write_heat_legend(ws, start_row + 1, 1)

    header_row = start_row + 2
    ws.cell(header_row, 1).value = "板块"
    ws.cell(header_row, 1).fill = section_fill
    ws.cell(header_row, 1).font = white_font
    ws.cell(header_row, 1).alignment = center
    apply_cell_border(ws.cell(header_row, 1), color="FFB7C9E2")

    for offset, (trade_date, _) in enumerate(recent_items, start=2):
        c = ws.cell(header_row, offset)
        c.value = datetime.strptime(trade_date, "%Y-%m-%d")
        c.number_format = "mm-dd"
        c.fill = section_fill
        c.font = white_font
        c.alignment = center
        apply_cell_border(c, color="FFB7C9E2")

    total_col = len(recent_items) + 2
    peak_col = total_col + 1
    ws.cell(header_row, total_col).value = "60日合计"
    ws.cell(header_row, peak_col).value = "单日峰值"
    for col_idx in (total_col, peak_col):
        c = ws.cell(header_row, col_idx)
        c.fill = section_fill
        c.font = white_font
        c.alignment = center
        apply_cell_border(c, color="FFB7C9E2")

    for row_offset, sector in enumerate(sector_order, start=1):
        current_row = header_row + row_offset
        label_cell = ws.cell(current_row, 1)
        label_cell.value = sector
        label_cell.fill = header_fill if row_offset % 2 else stripe_fill
        label_cell.font = strong_font
        label_cell.alignment = center
        apply_cell_border(label_cell, color="FFE2E8F0")

        main_row = row_map[sector]
        values = []
        for col_offset, (_, main_col) in enumerate(recent_items, start=2):
            count = parse_leading_int(ws_main.cell(main_row, main_col).value) or 0
            values.append(count)
            heat_cell = ws.cell(current_row, col_offset)
            heat_cell.value = count
            fill_color, font_color = build_heat_fill(count)
            heat_cell.fill = PatternFill("solid", fgColor=fill_color)
            heat_cell.font = Font(color=font_color, bold=(count >= 4))
            heat_cell.alignment = center
            apply_cell_border(heat_cell, color="FFFFFFFF")

        total_cell = ws.cell(current_row, total_col)
        peak_cell = ws.cell(current_row, peak_col)
        total_cell.value = sum(values)
        peak_cell.value = max(values) if values else 0
        for metric_cell in (total_cell, peak_cell):
            metric_cell.fill = PatternFill("solid", fgColor="FFF8FBFF")
            metric_cell.font = strong_font
            metric_cell.alignment = center
            apply_cell_border(metric_cell, color="FFE2E8F0")

    bottom_row = header_row + len(sector_order)

    # 列底部增加每日热度合计，更容易看出情绪高低切换
    footer_row = bottom_row + 1
    ws.cell(footer_row, 1).value = "单日合计"
    ws.cell(footer_row, 1).fill = section_fill
    ws.cell(footer_row, 1).font = white_font
    ws.cell(footer_row, 1).alignment = center
    apply_cell_border(ws.cell(footer_row, 1), color="FFB7C9E2")

    for col_idx in range(2, total_col):
        col_sum = sum((ws.cell(row_idx, col_idx).value or 0) for row_idx in range(header_row + 1, bottom_row + 1))
        c = ws.cell(footer_row, col_idx)
        c.value = col_sum
        fill_color, font_color = build_heat_fill(int(math.ceil(col_sum / max(len(sector_order), 1))))
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.font = Font(color=font_color, bold=True)
        c.alignment = center
        apply_cell_border(c, color="FFE2E8F0")

    for col_idx in (total_col, peak_col):
        c = ws.cell(footer_row, col_idx)
        c.value = None
        c.fill = PatternFill("solid", fgColor="FFEAF2FF")
        apply_cell_border(c, color="FFE2E8F0")

    # 列宽与行高
    ws.column_dimensions["A"].width = 16
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(7, 17):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["R"].width = 18
    ws.column_dimensions["S"].width = 10
    for col in range(2, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 5.6
    ws.column_dimensions[get_column_letter(total_col)].width = 10
    ws.column_dimensions[get_column_letter(peak_col)].width = 10

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[summary_header_row].height = 22
    ws.row_dimensions[header_row].height = 22
    for row_idx in range(header_row + 1, footer_row + 1):
        ws.row_dimensions[row_idx].height = 20


def determine_update_range(
    ws,
    user_start: Optional[str],
    user_end: Optional[str],
    mode: str,
    existing: Optional[Dict[str, int]] = None,
) -> Tuple[str, str]:
    today = datetime.today().strftime("%Y-%m-%d")
    existing = existing if existing is not None else get_existing_date_map(ws)

    if mode == "rebuild":
        if not user_start:
            raise SystemExit("rebuild 模式必须提供 --start-date，例如：--start-date 2026-02-01")
        return user_start, user_end or today

    if mode == "dashboard":
        return "", ""

    if existing:
        last_date = max(existing)
        start_date = (datetime.strptime(last_date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        start_date = user_start or (datetime.today() - timedelta(days=DEFAULT_RECENT_LOOKBACK_DAYS)).strftime("%Y-%m-%d")

    if user_start and user_start > start_date:
        start_date = user_start
    end_date = user_end or today
    return start_date, end_date


def validate_date_string(value: Optional[str], field_name: str) -> Optional[str]:
    if not value:
        return value
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as exc:
        raise SystemExit(f"{field_name} 格式错误，应为 YYYY-MM-DD，例如 2026-02-27") from exc


def save_workbook(wb: Workbook, output_path: Path) -> None:
    wb.save(output_path)
    logging.info("已保存：%s", output_path)


def resolve_output_path(args) -> Path:
    stats_path = Path(args.stats)
    return Path(args.output) if args.output else stats_path.with_name(f"{stats_path.stem}_updated{stats_path.suffix}")


def resolve_log_path(args, output_path: Path) -> Path:
    if getattr(args, "log_file", ""):
        return Path(args.log_file)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_path.with_name(f"{output_path.stem}_{args.mode}_{timestamp}.log")


def log_run_summary(summary: RunSummary) -> None:
    logging.info("======== 运行摘要 ========")
    logging.info("模式      ：%s", summary.mode)
    if summary.start_date or summary.end_date:
        logging.info("统计区间  ：%s ~ %s", summary.start_date, summary.end_date)
    logging.info("计划交易日：%s", len(summary.trade_dates))
    logging.info("成功交易日：%s", len(summary.success_dates))
    logging.info("失败交易日：%s", len(summary.failed_dates))
    if summary.success_dates:
        logging.info("成功范围  ：%s ~ %s", summary.success_dates[0], summary.success_dates[-1])
    if summary.failed_dates:
        logging.warning("失败日期  ：%s", ", ".join(summary.failed_dates))
    logging.info("输出文件  ：%s", summary.output_path)
    if summary.log_path:
        logging.info("日志文件  ：%s", summary.log_path)
    logging.info("==========================")


def run_update(args) -> RunSummary:
    configure_proxy_mode(args.proxy_mode)

    if args.mode != "dashboard":
        ensure_akshare()

    industry_path = Path(args.industry)
    stats_path = Path(args.stats)
    output_path = resolve_output_path(args)

    sector_order, sector_to_codes, code_to_sector = load_industry_mapping(industry_path)
    wb, ws = load_or_create_workbook(stats_path, sector_order, sector_to_codes)
    row_map = build_sector_row_map(ws, sector_order, sector_to_codes)
    existing_date_map = get_existing_date_map(ws, warn=True)

    if args.mode == "dashboard":
        rebuild_dashboard(wb, ws, sector_order, row_map)
        save_workbook(wb, output_path)
        return RunSummary(
            mode=args.mode,
            start_date="",
            end_date="",
            trade_dates=[],
            success_dates=[],
            failed_dates=[],
            output_path=output_path,
            log_path=getattr(args, "_resolved_log_path", None),
        )

    start_date, end_date = determine_update_range(ws, args.start_date, args.end_date, args.mode, existing=existing_date_map)
    start_date = validate_date_string(start_date, "start-date")
    end_date = validate_date_string(end_date, "end-date")

    if start_date > end_date:
        raise SystemExit("start-date 不能晚于 end-date")

    index_fetch_start = previous_calendar_day(start_date, 20)
    index_df = fetch_index_history(index_fetch_start, end_date, retries=args.retries, sleep_seconds=args.sleep, proxy_mode=args.proxy_mode)
    trade_dates = collect_trade_dates(index_df, start_date, end_date)
    if not trade_dates:
        logging.info("指定区间内没有可更新的交易日：%s ~ %s", start_date, end_date)
        rebuild_dashboard(wb, ws, sector_order, row_map)
        save_workbook(wb, output_path)
        return RunSummary(
            mode=args.mode,
            start_date=start_date,
            end_date=end_date,
            trade_dates=[],
            success_dates=[],
            failed_dates=[],
            output_path=output_path,
            log_path=getattr(args, "_resolved_log_path", None),
        )

    logging.info("待处理交易日 %s 个：%s ~ %s", len(trade_dates), trade_dates[0], trade_dates[-1])

    index_map = {row["日期字符串"]: row for _, row in index_df.iterrows()}
    success_dates: List[str] = []
    failed_dates: List[str] = []

    def process_one_trade_date(trade_date: str, retries: int, sleep_seconds: float) -> None:
        pool_df = fetch_limit_up_pool(trade_date, retries=retries, sleep_seconds=sleep_seconds, proxy_mode=args.proxy_mode)
        result = build_daily_result(trade_date, pool_df, index_map[trade_date], sector_order, code_to_sector)
        col = ensure_date_column(ws, trade_date, len(sector_order), existing_date_map=existing_date_map)
        write_daily_result(ws, col, result, sector_order, row_map)

    for trade_date in trade_dates:
        try:
            process_one_trade_date(trade_date, retries=args.retries, sleep_seconds=args.sleep)
            success_dates.append(trade_date)
            time.sleep(max(args.sleep, 0.2))
        except Exception as exc:  # pragma: no cover
            failed_dates.append(trade_date)
            logging.error("处理 %s 失败：%s", trade_date, exc)

    if failed_dates:
        retry_candidates = failed_dates[:]
        failed_dates = []
        logging.warning("首次处理后仍有 %s 个交易日失败，开始二次补抓。", len(retry_candidates))
        for trade_date in retry_candidates:
            try:
                process_one_trade_date(
                    trade_date,
                    retries=max(args.retries + 1, 4),
                    sleep_seconds=max(args.sleep * 1.5, 1.2),
                )
                success_dates.append(trade_date)
                logging.info("二次补抓成功：%s", trade_date)
                time.sleep(max(args.sleep, 0.2))
            except Exception as exc:  # pragma: no cover
                failed_dates.append(trade_date)
                logging.error("二次补抓仍失败 %s：%s", trade_date, exc)

    rebuild_dashboard(wb, ws, sector_order, row_map)
    save_workbook(wb, output_path)

    success_dates = sorted(set(success_dates))
    failed_dates = sorted(set(failed_dates))
    logging.info("完成：成功 %s 个交易日，失败 %s 个交易日。", len(success_dates), len(failed_dates))

    return RunSummary(
        mode=args.mode,
        start_date=start_date,
        end_date=end_date,
        trade_dates=trade_dates,
        success_dates=success_dates,
        failed_dates=failed_dates,
        output_path=output_path,
        log_path=getattr(args, "_resolved_log_path", None),
    )


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="按 28 个行业板块统计每日涨停数并输出 Excel")
    parser.add_argument("--industry", required=True, help="TDX_Industry_merged.xlsx 路径")
    parser.add_argument("--stats", required=True, help="历史统计 Excel 路径；不存在时会创建新文件")
    parser.add_argument("--output", default="", help="输出 Excel 路径；不填则自动生成 *_updated.xlsx")
    parser.add_argument("--mode", choices=["update", "rebuild", "dashboard"], default="update", help="update=增量更新；rebuild=指定区间重算；dashboard=仅刷新图表")
    parser.add_argument("--start-date", default="", help="开始日期，格式 YYYY-MM-DD")
    parser.add_argument("--end-date", default="", help="结束日期，格式 YYYY-MM-DD")
    parser.add_argument("--sleep", type=float, default=1.0, help="每次网络请求后的基础等待秒数，默认 1.0")
    parser.add_argument("--retries", type=int, default=4, help="网络请求重试次数，默认 4")
    parser.add_argument("--log-file", default="", help="日志文件路径；不填则自动按输出文件名生成")
    parser.add_argument("--proxy-mode", choices=["auto", "direct"], default="auto", help="网络代理模式：auto=默认读取系统/环境代理，若检测到代理异常则自动切直连；direct=启动时直接禁用代理")
    parser.add_argument("--verbose", action="store_true", help="输出更详细日志")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    output_path = resolve_output_path(args)
    args._resolved_log_path = resolve_log_path(args, output_path)
    setup_logging(args.verbose, args._resolved_log_path)

    try:
        summary = run_update(args)
        log_run_summary(summary)
        logging.info("脚本执行结束。")
        return 0
    except KeyboardInterrupt:  # pragma: no cover
        logging.error("用户中断执行。")
        return 130
    except Exception as exc:  # pragma: no cover
        logging.exception("执行失败：%s", exc)
        logging.error("如需排查，请查看日志文件：%s", args._resolved_log_path)
        return 1


if __name__ == "__main__":
    sys.exit(main())
